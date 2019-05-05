# encoding=utf_8
# @Author   ： 豆子
# @Function :  excel文件解析


from openpyxl import load_workbook


class Cases:
    def __init__(self):
        self.sheet_name = None
        self.case_id = None
        self.url = None
        self.method = None
        self.data = None
        self.title = None
        self.expected = None
        self.actual = None
        self.result = None
        # self.sql = None


class ExcelParse(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename)

    def get_cases(self, sheetname):
        '''
        :param sheetname: excel工作表的名称
        :param array: 读取工作表的用例id，需要传入一个list，默认读取所有
        :return: 返回读取的用例
        '''
        sheet = self.wb[sheetname]
        all_cases = []
        for row in range(2, sheet.max_row+1):
            row_case = Cases()
            row_case.sheet_name = sheetname
            row_case.case_id = sheet.cell(row, 1).value
            row_case.title = sheet.cell(row, 2).value
            row_case.url = sheet.cell(row, 3).value
            row_case.method = sheet.cell(row, 4).value
            row_case.data = sheet.cell(row, 5).value
            row_case.expected = sheet.cell(row, 6).value
            # row_case.sql = sheet.cell(row, 9).value
            all_cases.append(row_case)

        self.wb.close()
        return all_cases

    def back_write_by_excel(self, sheetname, case_id, actual, result):
        '''
        :param sheetname: excel工作表的名称
        :param case_id: 写入用例的id
        :param actual: 服务器返回的数据
        :param result: 用例的执行结果
        :return:
        '''
        sheet = self.wb[sheetname]
        sheet.cell(row=case_id+1, column=7, value=actual)
        sheet.cell(row=case_id+1, column=8, value=result)
        self.wb.save(self.filename)
        self.wb.close()


